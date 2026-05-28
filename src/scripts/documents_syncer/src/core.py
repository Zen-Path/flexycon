import csv
import hashlib
import io
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from common.cmd_utilities import run_cmd
from common.logger import log

STORE_FILE = Path("microsoft_documents_store.json")

SUPPORTED_EXTENSIONS = {
    ".docx": ".html",
    ".pdf": ".txt",
    ".xlsx": ".csv",
}

# Directories we completely ignore during scanning
IGNORED_DIRS = {".git", "node_modules", ".venv"}


def get_file_hash(filepath: Path) -> str:
    """Returns the SHA-256 hash of a file."""
    if not filepath.exists():
        return ""

    hasher = hashlib.sha256()
    with open(filepath, "rb") as f:
        while chunk := f.read(8192):
            hasher.update(chunk)

    return hasher.hexdigest()


def load_store(store_path: Path | None = None) -> dict[str, Any]:
    """Loads the JSON store and returns a dictionary keyed by source_path."""
    store_file = store_path or STORE_FILE

    try:
        with open(store_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            # Convert list of dicts to a dict keyed by source_path for easy lookup
            return {item["source_path"]: item for item in data}

    except FileNotFoundError:
        log.warning(f"File  {str(store_file)!r} not found. Starting fresh.")

    except json.JSONDecodeError:
        log.warning(f"File {str(store_file)!r} is corrupted. Starting fresh.")

    return {}


def save_store(store_dict: dict[str, Any], store_path: Path | None = None):
    """Saves the dictionary back to the JSON file as a list of dicts."""
    store_file = store_path or STORE_FILE

    # Sort the dictionary values alphabetically by the 'source_path' key
    sorted_data = sorted(store_dict.values(), key=lambda item: item["source_path"])

    with open(store_file, "w", encoding="utf-8") as f:
        json.dump(sorted_data, f, indent=4)

    run_cmd(["prettier", store_file, "-w"])


def get_destination_path(source_file: Path) -> Path | None:
    """Calculates the destination path based on the source file's extension."""
    target_suffix = SUPPORTED_EXTENSIONS.get(source_file.suffix)
    if not target_suffix:
        return None

    return source_file.with_suffix(target_suffix)


def is_ignored(path: Path) -> bool:
    """Returns True if the path belongs to an ignored directory or temporary file."""
    # Check if any parent part belongs to an ignored directory
    if any(part in IGNORED_DIRS for part in path.parts):
        return True

    # Check for Microsoft temporary files
    if path.name.startswith("~$"):
        return True

    return False


def collect_target_files() -> list[Path]:
    """Collects valid files for conversion via a workspace scan."""
    valid_files: list[Path] = []

    for ext in SUPPORTED_EXTENSIONS:
        for path in Path(".").rglob(f"*{ext}"):
            if path.is_file() and not is_ignored(path):
                valid_files.append(path)

    return valid_files


# CONVERTERS


def convert_docx(source_file: Path, dest_file: Path) -> tuple[bool, str | None]:
    """
    Convert a Microsoft Word (docx) file to HTML.

    Returns:
        tuple of (success, error_msg)
    """

    pandoc_result = run_cmd(
        ["pandoc", source_file, "-t", "html"],
    )
    if not pandoc_result.success:
        return False, f"Conversion with 'pandoc' failed.\n{pandoc_result.output}"

    with open(dest_file, "w", encoding="utf-8") as f:
        f.write(pandoc_result.output)

    # Format destination contents
    prettier_result = run_cmd(
        ["prettier", "--parser", "html", dest_file, "--write"],
    )
    if not prettier_result.success:
        return (False, f"Formatting with 'prettier' failed.\n{prettier_result.output}")

    return True, None


def convert_excel(source_file: Path, dest_file: Path) -> tuple[bool, str | None]:
    """
    Convert all sheets of an Excel file into a single consolidated CSV file.
    """

    wb: Workbook = openpyxl.load_workbook(filename=str(source_file), data_only=True)
    combined_output: list[str] = []

    sheet_name: str
    for sheet_name in wb.sheetnames:
        sheet: Worksheet = wb[sheet_name]

        # Read all rows into memory to perform boundary analysis
        raw_data: list[tuple[Any, ...]] = list(sheet.iter_rows(values_only=True))

        # Initialize global boundary trackers
        min_r: int = -1
        max_r: int = -1
        min_c: int = -1
        max_c: int = -1

        for r_idx, row in enumerate(raw_data):
            for c_idx, val in enumerate(row):
                if val is not None and str(val).strip() != "":
                    if min_r == -1 or r_idx < min_r:
                        min_r = r_idx
                    if max_r == -1 or r_idx > max_r:
                        max_r = r_idx
                    if min_c == -1 or c_idx < min_c:
                        min_c = c_idx
                    if max_c == -1 or c_idx > max_c:
                        max_c = c_idx

        # Checking min_r alone is sufficient to know if the sheet had data
        if min_r == -1:
            continue

        # NOTE: code is NOT unreachable, analyzer is just confused.

        sheet_buffer: io.StringIO = io.StringIO()
        writer = csv.writer(sheet_buffer, lineterminator="\n")

        # Extract rows and columns strictly within the calculated bounding box
        for r_idx in range(min_r, max_r + 1):
            raw_row: tuple[Any, ...] = raw_data[r_idx]

            # Slice row from the absolute leftmost to the absolute rightmost data column
            sliced_row: tuple[Any, ...] = raw_row[min_c : max_c + 1]

            # Sanitize values into clean strings
            cleaned_row: list[str] = [
                ""
                if v is None
                else str(v).strip().replace("\n", " ").replace("\r", " ")
                for v in sliced_row
            ]
            writer.writerow(cleaned_row)

        sheet_csv_content: str = sheet_buffer.getvalue().strip()

        if sheet_csv_content:
            formatted_sheet: str = f"--- {sheet_name} ---\n{sheet_csv_content}"
            combined_output.append(formatted_sheet)

    if not combined_output:
        return False, "Excel file contains no data across all sheets."

    final_csv_content: str = "\n\n".join(combined_output) + "\n"

    with open(dest_file, "w", encoding="utf-8") as f:
        f.write(final_csv_content)

    return True, None


def convert_pdf(source_file: Path, dest_file: Path) -> tuple[bool, str | None]:
    """
    Convert PDF file to text.

    Returns:
        tuple of (success, error_msg)
    """

    result = run_cmd(["pdftotext", source_file, dest_file])
    if not result.success:
        return False, f"Conversion with 'pdftotext' failed.\n{result.output}"

    return True, None


def run_conversion(source_file: Path, dest_file: Path) -> bool:
    """Convert source file to destination based on the source's extension"""

    log.info(f"Processing: {str(source_file)!r} -> {str(dest_file.name)!r}")

    extension = source_file.suffix
    match extension:
        case ".docx":
            conversion_fn = convert_docx
        case ".pdf":
            conversion_fn = convert_pdf
        case ".xlsx":
            conversion_fn = convert_excel
        case _:
            log.warning(f"Unknown source file extension {extension!r}")
            return False

    try:
        success, error_msg = conversion_fn(source_file, dest_file)
        if success:
            return True

        log.error(f"Conversion failed for {str(source_file)!r}: {error_msg}")

    except Exception as e:
        log.error(f"Unable to convert {str(source_file)!r}: {e}")

    return False
